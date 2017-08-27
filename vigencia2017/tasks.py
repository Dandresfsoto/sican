#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from __future__ import absolute_import

from sican.celery import app
import openpyxl
from informes.functions import construir_reporte
from datetime import datetime
from usuarios.models import User
from django.core.files import File
import pytz
from productos.models import Diplomado
from region.models import Region
from formadores.models import Formador
from formadores.models import Grupos
from matrices.models import Beneficiario, Area, Grado
from radicados.models import Radicado
from sican.settings import base as settings
import PIL
from PIL import ImageFont
from PIL import Image
from PIL import ImageDraw
import StringIO
from django.core.files import File
from django.utils import timezone
from vigencia2017.models import CargaMatriz, DaneSEDE, Grupos, Beneficiario, BeneficiarioCambio
from formadores.models import Contrato
from validate_email import validate_email
from formadores.models import Contrato as ContratoVigencia2017
from vigencia2017.models import Grupos as GruposVigencia2017
from vigencia2017.models import Beneficiario as BeneficiarioVigencia2017
from vigencia2017.models import Evidencia as EvidenciaVigencia2017
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font
from informes.models import InformesExcel
from productos.models import Entregable

@app.task
def carga_masiva_matrices(id,email_user):
    print(id)
    print(email_user)
    carga = CargaMatriz.objects.get(id = id)
    wb = openpyxl.load_workbook(filename = carga.archivo.file.name,read_only=True)
    sheets = wb.get_sheet_names()

    if 'InnovaTIC' in sheets and 'TecnoTIC' in sheets and 'DirecTIC' in sheets:

        titulos = ['DIPLOMADO',
                   'RESULTADO',
                   'REGION',
                   'CODIGO DANE SEDE EDUCATIVA',
                   'NOMBRE DE LA SEDE EDUCATIVA',
                   'CODIGO DANE INSTITUCION EDUCATIVA',
                   'NOMBRE INSTITUCION EDUCATIVA',
                   'CODIGO MUNICIPIO',
                   'MUNICIPIO',
                   'CODIGO DEPARTAMENTO',
                   'DEPARTAMENTO',
                   'SECRETARIA DE EDUCACION',
                   'ZONA (URBANA/RURAL)',
                   'CODIGO DEL GRUPO',
                   'NOMBRE DEL FORMADOR',
                   'NUMERO DE CEDULA DEL FORMADOR',
                   'APELLIDOS DEL DOCENTE',
                   'NOMBRES DEL DOCENTE',
                   'NUMERO DE CEDULA DEL DOCENTE',
                   'CORREO ELECTRONICO',
                   'TELEFONO FIJO',
                   'TELEFONO CELULAR',
                   'AREA CURRICULAR',
                   'GRADO',
                   'TIPO DE BENEFICIARIO',
                   'GENERO']

        formatos = ['General',
                   'General',
                   'General',
                   '0',
                   'General',
                   '0',
                   'General',
                   '0',
                   'General',
                   '0',
                   'General',
                   'General',
                   'General',
                   'General',
                   'General',
                   '0',
                   'General',
                   'General',
                   '0',
                   'General',
                   'General',
                   'General',
                   '0',
                   '0',
                   'General',
                   'General']

        ancho_columnas =  [30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           30,
                           ]

        contenidos = []

        for name in ['InnovaTIC','TecnoTIC','DirecTIC']:
            ws = wb.get_sheet_by_name(name)

            for fila in ws.iter_rows(row_offset=5):

                resultado = ''


                try:
                    diplomado = Diplomado.objects.get(nombre__icontains = name)
                except:
                    resultado = 'No existe el diplomado'
                else:

                    try:
                        region = Region.objects.get(id = fila[0].value)
                    except:
                        resultado = 'Codigo invalido de región'
                    else:

                        try:
                            dane_sede = DaneSEDE.objects.get(dane_sede = fila[1].value)
                        except:
                            resultado = 'No existe el codigo DANE de la sede'
                        else:

                            ruta_archivo = fila[11].value.split('-')

                            if len(ruta_archivo) == 3:

                                ruta = ruta_archivo[0]+"-"+ruta_archivo[1]

                                try:
                                    formador = Formador.objects.get(cedula=fila[13].value)
                                except:
                                    resultado = 'No existe el numero de cedula del formador'
                                else:
                                    try:
                                        contrato = Contrato.objects.get(formador=formador, codigo_ruta = ruta)
                                    except:
                                        resultado = "No se pudo identificar el contrato del formador (contacte a sistemas)"
                                    else:
                                        try:
                                            numero_grupo = int(ruta_archivo[2])
                                        except:
                                            resultado = "Numero de grupo invalido"
                                        else:
                                            grupo, created = Grupos.objects.get_or_create(contrato = contrato,
                                                                                 diplomado = diplomado,
                                                                                 numero = numero_grupo)
                                            try:
                                                cedula = long(fila[16].value)
                                            except:
                                                resultado = "Error en el numero de cedula"
                                            else:

                                                if fila[14].value != None and fila[15].value != None:

                                                    nombres = fila[15].value.upper()
                                                    apellidos = fila[14].value.upper()

                                                    email = ''
                                                    telefono_fijo = fila[18].value
                                                    telefono_celular = fila[19].value

                                                    genero = fila[23].value

                                                    if validate_email(fila[17].value):
                                                        email = fila[17].value

                                                    try:
                                                        area = int(fila[20].value)
                                                    except:
                                                        area = None

                                                    try:
                                                        grado = int(fila[21].value)
                                                    except:
                                                        grado = None

                                                    try:
                                                        beneficiario = Beneficiario.objects.get(cedula = cedula)
                                                    except:
                                                        resultado = "Beneficiario creado"
                                                        Beneficiario.objects.create(region=region,
                                                                                    dane_sede=dane_sede,
                                                                                    grupo=grupo,
                                                                                    apellidos=apellidos,
                                                                                    nombres=nombres,
                                                                                    cedula=cedula,
                                                                                    correo=email,
                                                                                    telefono_fijo=telefono_fijo,
                                                                                    telefono_celular=telefono_celular,
                                                                                    area=area,
                                                                                    grado=grado,
                                                                                    genero=genero
                                                                                    )
                                                    else:
                                                        resultado = "Solicitud de cambio"
                                                        if beneficiario.nombres != nombres and beneficiario.apellidos != apellidos:
                                                            if beneficiario.dane_sede != dane_sede and beneficiario.grupo != grupo:

                                                                BeneficiarioCambio.objects.create(original = beneficiario,
                                                                                                  masivo = carga,
                                                                                                  region = region,
                                                                                                  dane_sede = dane_sede,
                                                                                                  grupo = grupo,
                                                                                                  apellidos = apellidos,
                                                                                                  nombres = nombres,
                                                                                                  cedula=cedula,
                                                                                                  correo=email,
                                                                                                  telefono_fijo=telefono_fijo,
                                                                                                  telefono_celular=telefono_celular,
                                                                                                  area=area,
                                                                                                  grado=grado,
                                                                                                  genero=genero
                                                                                                  )
                                                        else:
                                                            beneficiario.dane_sede = dane_sede
                                                            beneficiario.grupo = grupo
                                                            beneficiario.apellidos = apellidos
                                                            beneficiario.nombres = nombres
                                                            beneficiario.correo = email
                                                            beneficiario.telefono_fijo = telefono_fijo
                                                            beneficiario.telefono_celular = telefono_celular
                                                            beneficiario.area = area
                                                            beneficiario.grado = grado
                                                            beneficiario.genero = genero
                                                            beneficiario.save()

                                                else:
                                                    resultado = "Error en los nombres o apellidos del docente"


                            else:
                                resultado = 'El codigo de ruta no se encuentra bien parametrizado'






                contenidos.append([
                    name,
                    resultado,
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    fila[3].value,
                    fila[4].value,
                    fila[5].value,
                    fila[6].value,
                    fila[7].value,
                    fila[8].value,
                    fila[9].value,
                    fila[10].value,
                    fila[11].value,
                    fila[12].value,
                    fila[13].value,
                    fila[14].value,
                    fila[15].value,
                    fila[16].value,
                    fila[17].value,
                    fila[18].value,
                    fila[19].value,
                    fila[20].value,
                    fila[21].value,
                    fila[22].value,
                    fila[23].value
                ])

        usuario = User.objects.get(email=email_user)
        nombre = "Resultado carga masiva matrices"
        proceso = "FOR-MAS01"
        fecha = pytz.utc.localize(datetime.now())
        output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
        filename = unicode(fecha) + '.xlsx'
        carga.resultado.save(filename,File(output))

    elif 'Matriz revisión documental' in sheets:
        titulos = ['DIPLOMADO',
                   'RESULTADO',
                   'REGION',
                   'CODIGO DANE SEDE EDUCATIVA',
                   'NOMBRE DE LA SEDE EDUCATIVA',
                   'CODIGO DANE INSTITUCION EDUCATIVA',
                   'NOMBRE INSTITUCION EDUCATIVA',
                   'CODIGO MUNICIPIO',
                   'MUNICIPIO',
                   'CODIGO DEPARTAMENTO',
                   'DEPARTAMENTO',
                   'SECRETARIA DE EDUCACION',
                   'ZONA (URBANA/RURAL)',
                   'CODIGO DEL GRUPO',
                   'NOMBRE DEL FORMADOR',
                   'NUMERO DE CEDULA DEL FORMADOR',
                   'APELLIDOS DEL DOCENTE',
                   'NOMBRES DEL DOCENTE',
                   'NUMERO DE CEDULA DEL DOCENTE',
                   'CORREO ELECTRONICO',
                   'TELEFONO FIJO',
                   'TELEFONO CELULAR',
                   'AREA CURRICULAR',
                   'GRADO',
                   'TIPO DE BENEFICIARIO',
                   'GENERO']

        formatos = ['General',
                    'General',
                    'General',
                    '0',
                    'General',
                    '0',
                    'General',
                    '0',
                    'General',
                    '0',
                    'General',
                    'General',
                    'General',
                    'General',
                    'General',
                    '0',
                    'General',
                    'General',
                    '0',
                    'General',
                    'General',
                    'General',
                    '0',
                    '0',
                    'General',
                    'General']

        ancho_columnas = [30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          30,
                          ]

        contenidos = []

        for name in ['Matriz revisión documental']:
            ws = wb.get_sheet_by_name(name)

            for fila in ws.iter_rows(row_offset=9):

                resultado = ''

                try:
                    diplomado = Diplomado.objects.get(id = 4)
                except:
                    resultado = 'No existe el diplomado'
                else:

                    try:
                        region = Region.objects.get(id=fila[0].value)
                    except:
                        resultado = 'Codigo invalido de región'
                    else:

                        try:
                            dane_sede = None
                        except:
                            resultado = 'No existe el codigo DANE de la sede'
                        else:

                            ruta_archivo = fila[11].value.split('-')

                            if len(ruta_archivo) == 3:

                                ruta = ruta_archivo[0] + "-" + ruta_archivo[1]

                                try:
                                    formador = Formador.objects.get(cedula=fila[13].value)
                                except:
                                    resultado = 'No existe el numero de cedula del formador'
                                else:
                                    try:
                                        contrato = Contrato.objects.get(formador=formador, codigo_ruta=ruta)
                                    except:
                                        resultado = "No se pudo identificar el contrato del formador (contacte a sistemas)"
                                    else:
                                        try:
                                            numero_grupo = int(ruta_archivo[2])
                                        except:
                                            resultado = "Numero de grupo invalido"
                                        else:
                                            grupo, created = Grupos.objects.get_or_create(contrato=contrato,
                                                                                          diplomado=diplomado,
                                                                                          numero=numero_grupo)
                                            try:
                                                cedula = long(fila[16].value)
                                            except:
                                                resultado = "Error en el numero de cedula"
                                            else:

                                                if fila[14].value != None and fila[15].value != None:

                                                    nombres = fila[15].value.upper()
                                                    apellidos = fila[14].value.upper()

                                                    email = ''
                                                    telefono_fijo = fila[18].value
                                                    telefono_celular = fila[19].value

                                                    genero = fila[23].value

                                                    if validate_email(fila[17].value):
                                                        email = fila[17].value

                                                    try:
                                                        area = int(fila[20].value)
                                                    except:
                                                        area = None

                                                    try:
                                                        grado = int(fila[21].value)
                                                    except:
                                                        grado = None

                                                    try:
                                                        beneficiario = Beneficiario.objects.get(cedula=cedula)
                                                    except:
                                                        resultado = "Beneficiario creado"
                                                        Beneficiario.objects.create(region=region,
                                                                                    dane_sede=dane_sede,
                                                                                    grupo=grupo,
                                                                                    apellidos=apellidos,
                                                                                    nombres=nombres,
                                                                                    cedula=cedula,
                                                                                    correo=email,
                                                                                    telefono_fijo=telefono_fijo,
                                                                                    telefono_celular=telefono_celular,
                                                                                    area=area,
                                                                                    grado=grado,
                                                                                    genero=genero
                                                                                    )
                                                    else:
                                                        resultado = "Solicitud de cambio"
                                                        if beneficiario.nombres != nombres and beneficiario.apellidos != apellidos:
                                                            if beneficiario.grupo != grupo:
                                                                BeneficiarioCambio.objects.create(
                                                                    original=beneficiario,
                                                                    masivo=carga,
                                                                    region=region,
                                                                    dane_sede=dane_sede,
                                                                    grupo=grupo,
                                                                    apellidos=apellidos,
                                                                    nombres=nombres,
                                                                    cedula=cedula,
                                                                    correo=email,
                                                                    telefono_fijo=telefono_fijo,
                                                                    telefono_celular=telefono_celular,
                                                                    area=area,
                                                                    grado=grado,
                                                                    genero=genero
                                                                    )
                                                        else:
                                                            beneficiario.dane_sede = dane_sede
                                                            beneficiario.grupo = grupo
                                                            beneficiario.apellidos = apellidos
                                                            beneficiario.nombres = nombres
                                                            beneficiario.correo = email
                                                            beneficiario.telefono_fijo = telefono_fijo
                                                            beneficiario.telefono_celular = telefono_celular
                                                            beneficiario.area = area
                                                            beneficiario.grado = grado
                                                            beneficiario.genero = genero
                                                            beneficiario.save()

                                                else:
                                                    resultado = "Error en los nombres o apellidos del docente"


                            else:
                                resultado = 'El codigo de ruta no se encuentra bien parametrizado'

                contenidos.append([
                    name,
                    resultado,
                    fila[0].value,
                    fila[1].value,
                    fila[2].value,
                    fila[3].value,
                    fila[4].value,
                    fila[5].value,
                    fila[6].value,
                    fila[7].value,
                    fila[8].value,
                    fila[9].value,
                    fila[10].value,
                    fila[11].value,
                    fila[12].value,
                    fila[13].value,
                    fila[14].value,
                    fila[15].value,
                    fila[16].value,
                    fila[17].value,
                    fila[18].value,
                    fila[19].value,
                    fila[20].value,
                    fila[21].value,
                    fila[22].value,
                    fila[23].value
                ])

        usuario = User.objects.get(email=email_user)
        nombre = "Resultado carga masiva matrices"
        proceso = "FOR-MAS01"
        fecha = pytz.utc.localize(datetime.now())
        output = construir_reporte(titulos, contenidos, formatos, ancho_columnas, nombre, fecha, usuario, proceso)
        filename = unicode(fecha) + '.xlsx'
        carga.resultado.save(filename, File(output))


    return "Carga masiva exitosa"

@app.task
def matriz_chequeo_vigencia_2017(email,id_contrato):
    usuario = User.objects.get(email=email)

    contrato = ContratoVigencia2017.objects.get(id = int(id_contrato))
    id_diplomado = ''
    nombre = ''


    grupos = GruposVigencia2017.objects.filter(contrato = contrato)


    proceso = "REV-INF06"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    informe.nombre = "Matriz de chequeo - " + contrato.formador.get_full_name() + " - " +contrato.nombre
    informe.save()
    fecha = informe.creacion
    output = StringIO.StringIO()
    dict_productos = []


    wb = openpyxl.load_workbook(filename = settings.STATICFILES_DIRS[0]+'/documentos/chequeo_2017.xlsx')
    ws_innovatic = wb.get_sheet_by_name('InnovaTIC')
    ws_tecnotic = wb.get_sheet_by_name('TecnoTIC')
    ws_directic = wb.get_sheet_by_name('DirecTIC')
    ws_escuelatic = wb.get_sheet_by_name('EscuelaTIC')


    contadores = {'1':6,'2':6,'3':6,'4':6}


    number = Style(font=Font(name='Calibri', size=12),
                   alignment=Alignment(horizontal='right', vertical='center', wrap_text=False),
                   number_format='0',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                 bottom=Side(style='thin'))
                   )

    text = Style(font=Font(name='Calibri', size=12),
                 alignment=Alignment(horizontal='left', vertical='center', wrap_text=False),
                 number_format='General',
                 border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                               bottom=Side(style='thin'))
                 )

    validado = Style(font=Font(name='Calibri', size=12),
                     alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                     number_format='General',
                     fill=PatternFill(fill_type='solid', start_color='FF00B050', end_color='FF00B050')
                     )

    enviado = Style(font=Font(name='Calibri', size=12),
                    alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                    number_format='General',
                    fill=PatternFill(fill_type='solid', start_color='FFFFC000', end_color='FFFFC000')
                    )

    cargado = Style(font=Font(name='Calibri', size=12),
                    alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                    number_format='General',
                    border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                                  bottom=Side(style='thin')),
                    )

    rechazado = Style(font=Font(name='Calibri', size=12),
                      alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                      number_format='General',
                      fill=PatternFill(fill_type='solid', start_color='FFFF0000', end_color='FFFF0000')
                      )



    for beneficiario in BeneficiarioVigencia2017.objects.filter(grupo__in = grupos):

        id_diplomado = beneficiario.grupo.diplomado.id

        if id_diplomado == 1:
            ws = ws_innovatic
        elif id_diplomado == 2:
            ws = ws_tecnotic
        elif id_diplomado == 3:
            ws = ws_directic
        elif id_diplomado == 4:
            ws = ws_escuelatic

        ws.cell(row=contadores[str(id_diplomado)], column=1, value = beneficiario.region.nombre.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=2, value = beneficiario.dane_sede.dane_sede if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=3, value = beneficiario.dane_sede.nombre_sede.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=4, value = beneficiario.dane_sede.dane_ie if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=5, value = beneficiario.dane_sede.nombre_ie.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=6, value = beneficiario.dane_sede.municipio.codigo_municipio if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=7, value = beneficiario.dane_sede.municipio.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=8, value = beneficiario.dane_sede.municipio.departamento.codigo_departamento if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=9, value = beneficiario.dane_sede.municipio.departamento.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=10, value = beneficiario.dane_sede.secretaria.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=11, value = beneficiario.dane_sede.zona.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=12, value = beneficiario.grupo.get_nombre_grupo())
        ws.cell(row=contadores[str(id_diplomado)], column=13, value = beneficiario.grupo.contrato.formador.get_full_name().upper())
        ws.cell(row=contadores[str(id_diplomado)], column=14, value = beneficiario.grupo.contrato.formador.cedula)
        ws.cell(row=contadores[str(id_diplomado)], column=15, value = beneficiario.apellidos.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=16, value = beneficiario.nombres.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=17, value = beneficiario.cedula)
        ws.cell(row=contadores[str(id_diplomado)], column=18, value = beneficiario.correo.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=19, value = beneficiario.telefono_fijo.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=20, value = beneficiario.telefono_celular.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=21, value = beneficiario.area)
        ws.cell(row=contadores[str(id_diplomado)], column=22, value = beneficiario.grado)
        ws.cell(row=contadores[str(id_diplomado)], column=23, value = beneficiario.grupo.diplomado.nombre.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=24, value = beneficiario.genero.upper())

        for entregable in Entregable.objects.filter(sesion__nivel__diplomado__id = id_diplomado):

            estado = beneficiario.get_evidencia_state(id_entregable = entregable.id)

            if estado['state'] == 'cargado':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value="C")
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = enviado
            elif estado['state'] == 'validado':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value="V")
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = validado
            elif estado['state'] == 'rechazado':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value="R")
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = rechazado

        contadores[str(id_diplomado)] += 1


    if contadores['1'] == 6:
        wb.remove_sheet(ws_innovatic)
    if contadores['2'] == 6:
        wb.remove_sheet(ws_tecnotic)
    if contadores['3'] == 6:
        wb.remove_sheet(ws_directic)
    if contadores['4'] == 6:
        wb.remove_sheet(ws_escuelatic)

    wb.save(output)

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def matriz_valores_vigencia_2017(email,id_contrato):
    usuario = User.objects.get(email=email)

    contrato = ContratoVigencia2017.objects.get(id = int(id_contrato))
    id_diplomado = ''
    nombre = ''


    grupos = GruposVigencia2017.objects.filter(contrato = contrato)


    proceso = "REV-INF06"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    informe.nombre = "Matriz de pago - " + contrato.formador.get_full_name() + " - " +contrato.nombre
    informe.save()
    fecha = informe.creacion
    output = StringIO.StringIO()
    dict_productos = []


    wb = openpyxl.load_workbook(filename = settings.STATICFILES_DIRS[0]+'/documentos/chequeo_2017.xlsx')
    ws_innovatic = wb.get_sheet_by_name('InnovaTIC')
    ws_tecnotic = wb.get_sheet_by_name('TecnoTIC')
    ws_directic = wb.get_sheet_by_name('DirecTIC')
    ws_escuelatic = wb.get_sheet_by_name('EscuelaTIC')


    contadores = {'1':6,'2':6,'3':6,'4':6}




    validado = Style(font=Font(name='Calibri', size=12),
                     alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                     number_format='$ #,##0.00',
                     fill=PatternFill(fill_type='solid', start_color='FF00B050', end_color='FF00B050')
                     )

    enviado = Style(font=Font(name='Calibri', size=12),
                    alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                    number_format='$ #,##0.00',
                    fill=PatternFill(fill_type='solid', start_color='FFFFC000', end_color='FFFFC000')
                    )


    rechazado = Style(font=Font(name='Calibri', size=12),
                      alignment=Alignment(horizontal='center', vertical='center', wrap_text=False),
                      number_format='$ #,##0.00',
                      fill=PatternFill(fill_type='solid', start_color='FFFF0000', end_color='FFFF0000')
                      )



    for beneficiario in BeneficiarioVigencia2017.objects.filter(grupo__in = grupos):

        id_diplomado = beneficiario.grupo.diplomado.id

        if id_diplomado == 1:
            ws = ws_innovatic
        elif id_diplomado == 2:
            ws = ws_tecnotic
        elif id_diplomado == 3:
            ws = ws_directic
        elif id_diplomado == 4:
            ws = ws_escuelatic

        ws.cell(row=contadores[str(id_diplomado)], column=1, value = beneficiario.region.nombre.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=2, value = beneficiario.dane_sede.dane_sede if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=3, value = beneficiario.dane_sede.nombre_sede.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=4, value = beneficiario.dane_sede.dane_ie if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=5, value = beneficiario.dane_sede.nombre_ie.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=6, value = beneficiario.dane_sede.municipio.codigo_municipio if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=7, value = beneficiario.dane_sede.municipio.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=8, value = beneficiario.dane_sede.municipio.departamento.codigo_departamento if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=9, value = beneficiario.dane_sede.municipio.departamento.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=10, value = beneficiario.dane_sede.secretaria.nombre.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=11, value = beneficiario.dane_sede.zona.upper() if beneficiario.dane_sede != None else "N/A")
        ws.cell(row=contadores[str(id_diplomado)], column=12, value = beneficiario.grupo.get_nombre_grupo())
        ws.cell(row=contadores[str(id_diplomado)], column=13, value = beneficiario.grupo.contrato.formador.get_full_name().upper())
        ws.cell(row=contadores[str(id_diplomado)], column=14, value = beneficiario.grupo.contrato.formador.cedula)
        ws.cell(row=contadores[str(id_diplomado)], column=15, value = beneficiario.apellidos.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=16, value = beneficiario.nombres.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=17, value = beneficiario.cedula)
        ws.cell(row=contadores[str(id_diplomado)], column=18, value = beneficiario.correo.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=19, value = beneficiario.telefono_fijo.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=20, value = beneficiario.telefono_celular.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=21, value = beneficiario.area)
        ws.cell(row=contadores[str(id_diplomado)], column=22, value = beneficiario.grado)
        ws.cell(row=contadores[str(id_diplomado)], column=23, value = beneficiario.grupo.diplomado.nombre.upper())
        ws.cell(row=contadores[str(id_diplomado)], column=24, value = beneficiario.genero.upper())

        for entregable in Entregable.objects.filter(sesion__nivel__diplomado__id = id_diplomado):

            estado = beneficiario.get_pago_state(id_entregable = entregable.id)

            if estado['state'] == 'reportado':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value = beneficiario.get_pago_valor_entregable(entregable.id))
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = enviado
            elif estado['state'] == 'pago':
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero), value = beneficiario.get_pago_valor_entregable(entregable.id))
                ws.cell(row=contadores[str(id_diplomado)], column=25 + int(entregable.numero)).style = validado
        contadores[str(id_diplomado)] += 1


    if contadores['1'] == 6:
        wb.remove_sheet(ws_innovatic)
    if contadores['2'] == 6:
        wb.remove_sheet(ws_tecnotic)
    if contadores['3'] == 6:
        wb.remove_sheet(ws_directic)
    if contadores['4'] == 6:
        wb.remove_sheet(ws_escuelatic)

    wb.save(output)

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"