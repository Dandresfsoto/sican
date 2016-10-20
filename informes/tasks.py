#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from __future__ import absolute_import

from sican.celery import app
from informes.functions import construir_reporte, cronograma_interventoria
from informes.models import InformesExcel
from django.core.files import File
from usuarios.models import User
from formadores.models import Formador, SolicitudTransporte
from formadores.models import Soporte as SoporteFormadores
from rh.models import TipoSoporte
from preinscripcion.models import DocentesPreinscritos
from formacion.models import EntradaCronograma
from isoweek import Week
import datetime
from formacion.models import Semana
from lideres.models import Lideres, Soporte
from encuestas.models import PercepcionInicial
from radicados.models import Radicado
from formadores.models import Cortes
from formadores.models import Revision
import zipfile
import shutil
import os
from rh.models import RequerimientoPersonal
from productos.models import Contratos, ValorEntregable
from StringIO import StringIO
import openpyxl
from sican.settings import base as settings
from productos.models import Diplomado
from evidencias.models import Beneficiario
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font
from evidencias.models import Evidencia
from productos.models import Entregable
from evidencias.models import Red
from openpyxl.comments import Comment
from django.db.models import Q

@app.task
def nueva_semana():
    x, created = Semana.objects.get_or_create(numero = datetime.datetime.now().isocalendar()[1]+1)
    return "Semana actualizada"

@app.task
def formadores(email):
    usuario = User.objects.get(email=email)
    nombre = "Directorio de formadores"
    proceso = "RH-INF01"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Nombres','Apellidos','Cedula','Región','Correo','Celular','Cargo','Profesión','Inicio contrato',
               'Fin contrato','Banco','Tipo cuenta','Numero cuenta','Eps','Pensión','Arl']

    formatos = ['General','General','General','General','General','General','General','General','General','d/m/yy',
               'd/m/yy','General','General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15,50,25,20,20,10,
                       10,20,20,20,20,20,20]

    contenidos = []

    for formador in Formador.objects.exclude(oculto=True):
        contenidos.append([
            'FOR-'+unicode(formador.id),
            formador.nombres,
            formador.apellidos,
            formador.cedula,
            formador.get_region_string(),
            formador.correo_personal,
            formador.celular_personal,
            formador.cargo.nombre if formador.cargo != None else '',
            formador.profesion,
            formador.fecha_contratacion,
            formador.fecha_terminacion,
            formador.banco.nombre if formador.banco != None else '',
            formador.tipo_cuenta,
            formador.numero_cuenta,
            formador.eps,
            formador.pension,
            formador.arl
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def reporte_requerimientos_contratacion(email):
    usuario = User.objects.get(email=email)
    nombre = "Requerimientos de contratacion"
    proceso = "RH-INF05"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Fecha Solicitud','Solicitante','Departamento','Municipios','Codigo Ruta','Encargado','Observación Solicitante',
               'Fecha Respuesta','Nombre','Cedula','Celular','Email','Hoja de vida','Observación respuesta','Fecha solicitud de contratación',
               'Observacion final','Estado']

    formatos = ['General','d/m/yy','General','General','General','General','General','General',
                'd/m/yy','General','0','General','General','General','General','d/m/yy',
                'General','General']


    ancho_columnas =  [30,15,20,15,30,15,20,30,
                       15,20,15,15,20,20,30,20,
                       30,30]

    contenidos = []

    for requerimiento in RequerimientoPersonal.objects.all():

        estado = ''

        if requerimiento.remitido_respuesta == False and requerimiento.remitido_contratacion == False and requerimiento.contratar == False and requerimiento.desierto == False and requerimiento.contratado == False and requerimiento.contrato_enviado == False:
            estado = 'Remitido a RH'

        elif requerimiento.remitido_respuesta == True and requerimiento.remitido_contratacion == False and requerimiento.contratar == False and requerimiento.desierto == False and requerimiento.contratado == False and requerimiento.contrato_enviado == False:
            estado = 'Listo para capacitar'

        elif requerimiento.remitido_respuesta == True and requerimiento.remitido_contratacion == True and requerimiento.contratar == True and requerimiento.desierto == False and requerimiento.contratado == False and requerimiento.contrato_enviado == False:
            estado = 'Proceder a contrato'

        elif requerimiento.remitido_respuesta == True and requerimiento.remitido_contratacion == True and requerimiento.contratar == False and requerimiento.desierto == True and requerimiento.contratado == False and requerimiento.contrato_enviado == False:
            estado = 'Aspirante deserta'

        elif requerimiento.contratado == False and requerimiento.contrato_enviado == True:
            estado = 'Contrato enviado'

        elif requerimiento.contratado == True:
            estado = 'Contratado'



        contenidos.append([
            'REQ-'+unicode(requerimiento.id),
            requerimiento.fecha_solicitud,
            requerimiento.solicitante.get_full_name_string(),
            requerimiento.departamento.nombre,
            requerimiento.get_municipios_string(),
            requerimiento.codigo_ruta,
            requerimiento.encargado.get_full_name_string(),
            requerimiento.observacion_solicitante,
            requerimiento.fecha_respuesta,
            requerimiento.nombre,
            requerimiento.cedula,
            requerimiento.celular,
            requerimiento.email,
            'Si' if str(requerimiento.hv) != '' else 'No',
            requerimiento.observacion_respuesta,
            requerimiento.fecha_solicitud_contratacion,
            requerimiento.observacion_final,
            estado
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def formadores_soportes(email):
    usuario = User.objects.get(email=email)
    nombre = "Soportes cargados por formador"

    proceso = "RH-INF02"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Nombres','Apellidos','Cedula','Región']

    formatos = ['General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15]

    contenidos = []

    tipos_soportes = TipoSoporte.objects.exclude(oculto=True).values_list('id','nombre')

    for tipo_soporte in tipos_soportes:
        titulos.append(tipo_soporte[1])
        formatos.append('General')
        ancho_columnas.append(30)

    for formador in Formador.objects.exclude(oculto=True):
        row =[
                'FOR-'+unicode(formador.id),
                formador.nombres,
                formador.apellidos,
                formador.cedula,
                formador.get_region_string(),
            ]
        for tipo_soporte in tipos_soportes:
            try:
                soporte = SoporteFormadores.objects.filter(formador=formador).get(tipo__id=tipo_soporte[0])
            except:
                row.append('No')
            else:
                if soporte.get_archivo_url() != "":
                    row.append('Si')
                else:
                    row.append('No')

        contenidos.append(row)

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def preinscritos(email):
    usuario = User.objects.get(email=email)
    nombre = "Reporte de docentes preinscritos"
    proceso = "FOR-INF01"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Cedula','Primer apellido','Segundo apellido','Primer nombre','Segundo nombre','Cargo','Correo',
               'Telefono fijo','Celular','Departamento','Municipio','Radicado','Base Mineducación','Fecha registro',
               'Secretaria Radicado','Municipio Radicado','Nombre Sede','Dane Sede','Tipo Sede','Ubicación']

    formatos = ['General','0','General','General','General','General','General','General',
                'General','General','General','General','0','General','d/m/yy',
                'General','General','General','General','General','General']


    ancho_columnas =  [30,20,15,15,15,15,15,30,
                       15,15,20,20,15,10,15,
                       20,20,20,20,20,20]

    contenidos = []

    for docente in DocentesPreinscritos.objects.all():
        contenidos.append([
            'PRE-'+unicode(docente.id),
            docente.cedula,
            docente.primer_apellido,
            docente.segundo_apellido,
            docente.primer_nombre,
            docente.segundo_nombre,
            docente.cargo,
            docente.correo,
            docente.telefono_fijo,
            docente.telefono_celular,
            docente.departamento.nombre if docente.departamento != None else '',
            docente.municipio.nombre if docente.municipio != None else '',
            docente.radicado.numero if docente.radicado != None else '',
            'Si' if docente.verificado else 'No',
            docente.fecha,
            docente.radicado.secretaria.nombre,
            docente.radicado.municipio.nombre,
            docente.radicado.nombre_sede,
            docente.radicado.dane_sede,
            docente.radicado.tipo,
            docente.radicado.ubicacion,
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def transportes(email):
    usuario = User.objects.get(email=email)
    nombre = "Reporte de solicitudes de desplazamiento"
    proceso = "FIN-INF01"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Formador','Cedula','Región','Cargo','Banco','Tipo de cuenta','Numero de cuenta',
               'Nombre solicitud','Fecha','Estado','Valor solicitado','Valor Aprobado','Observación']

    formatos = ['General','General','General','General','General','General','General','General',
                'General','d/m/yy h:mm:ss AM/PM','General','"$"#,##0_);("$"#,##0)','"$"#,##0_);("$"#,##0)','General']


    ancho_columnas =  [30,30,15,15,15,15,15,20,
                       50,15,20,15,15,60]

    contenidos = []

    for solicitud in SolicitudTransporte.objects.all():
        contenidos.append([
            'PRE-'+unicode(solicitud.id),
            solicitud.formador.get_full_name(),
            solicitud.formador.cedula,
            solicitud.formador.get_region_string(),
            solicitud.formador.cargo.nombre if solicitud.formador.cargo != None else "",
            solicitud.formador.banco.nombre if solicitud.formador.banco != None else "",
            solicitud.formador.tipo_cuenta,
            solicitud.formador.numero_cuenta,
            solicitud.nombre,
            solicitud.creacion,
            solicitud.estado,
            solicitud.valor,
            solicitud.valor_aprobado,
            solicitud.observacion
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def cronograma_general(email,semana_id):
    usuario = User.objects.get(email=email)
    nombre = "Cronograma consolidado general"

    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    semana = Semana.objects.get(id=semana_id)

    innovatics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 1").order_by('formador__region').values_list('id',flat=True)
    tecnotics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 2").order_by('formador__region').values_list('id',flat=True)
    directics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 3").order_by('formador__region').values_list('id',flat=True)
    escuelatics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 4").order_by('formador__region').values_list('id',flat=True)

    inicio = Week(semana.creacion.isocalendar()[0],semana.creacion.isocalendar()[1]+1).monday()
    fin = Week(semana.creacion.isocalendar()[0],semana.creacion.isocalendar()[1]+1).sunday()
    rango = inicio.strftime("Del dia %d de %B del %Y") + ' - ' + fin.strftime(" al dia %d de %B del %Y")

    output = cronograma_interventoria(innovatics,tecnotics,directics,escuelatics,rango)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def cronograma_lider(email,semana_id):
    usuario = User.objects.get(email=email)
    nombre = "Cronograma semanal lider"

    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    semana = Semana.objects.get(id=semana_id)

    innovatics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 1",formador__lider__email = email).values_list('id',flat=True)
    tecnotics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 2",formador__lider__email = email).values_list('id',flat=True)
    directics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 3",formador__lider__email = email).values_list('id',flat=True)
    escuelatics = EntradaCronograma.objects.filter(semana=semana,formador__cargo__nombre="Formador Tipo 4",formador__lider__email = email).values_list('id',flat=True)

    inicio = Week(semana.creacion.isocalendar()[0],semana.creacion.isocalendar()[1]+1).monday()
    fin = Week(semana.creacion.isocalendar()[0],semana.creacion.isocalendar()[1]+1).sunday()
    rango = inicio.strftime("Del dia %d de %B del %Y") + ' - ' + fin.strftime(" al dia %d de %B del %Y")

    output = cronograma_interventoria(innovatics,tecnotics,directics,escuelatics,rango)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def lideres(email):
    usuario = User.objects.get(email=email)
    nombre = "Directorio de lideres"
    proceso = "RH-INF03"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Nombres','Apellidos','Cedula','Región','Correo','Celular','Cargo','Profesión','Inicio contrato',
               'Fin contrato','Banco','Tipo cuenta','Numero cuenta','Eps','Pensión','Arl']

    formatos = ['General','General','General','General','General','General','General','General','General','d/m/yy',
               'd/m/yy','General','General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15,50,25,20,20,10,
                       10,20,20,20,20,20,20]

    contenidos = []

    for lider in Lideres.objects.exclude(oculto=True):
        contenidos.append([
            'LID-'+unicode(lider.id),
            lider.nombres,
            lider.apellidos,
            lider.cedula,
            lider.region.nombre,
            lider.correo_personal,
            lider.celular_personal,
            lider.cargo.nombre if lider.cargo != None else '',
            lider.profesion,
            lider.fecha_contratacion,
            lider.fecha_terminacion,
            lider.banco.nombre if lider.banco != None else '',
            lider.tipo_cuenta,
            lider.numero_cuenta,
            lider.eps,
            lider.pension,
            lider.arl
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def lideres_soportes(email):
    usuario = User.objects.get(email=email)
    nombre = "Soportes cargados por lider"
    proceso = "RH-INF04"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Nombres','Apellidos','Cedula','Región']

    formatos = ['General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15]

    contenidos = []

    tipos_soportes = TipoSoporte.objects.exclude(oculto=True).values_list('id','nombre')

    for tipo_soporte in tipos_soportes:
        titulos.append(tipo_soporte[1])
        formatos.append('General')
        ancho_columnas.append(30)

    for lider in Lideres.objects.exclude(oculto=True):
        row =[
                'LID-'+unicode(lider.id),
                lider.nombres,
                lider.apellidos,
                lider.cedula,
                lider.region.nombre,
            ]
        for tipo_soporte in tipos_soportes:
            try:
                soporte = Soporte.objects.filter(lider=lider).get(tipo__id=tipo_soporte[0])
            except:
                row.append('No')
            else:
                if soporte.get_archivo_url() != "":
                    row.append('Si')
                else:
                    row.append('No')

        contenidos.append(row)

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def encuesta_percepcion_inicial(email):
    usuario = User.objects.get(email=email)
    nombre = "Base de datos respuestas percepción inicial"
    proceso = "ENC-INF01"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Cedula','Primer Apellido','Segundo Apellido','Primer Nombre','Segundo Nombre','Cargo','Correo','Telefono Fijo','Telefono Celular',
               'Departamento','Municipio','Verificado','Fecha','Radicado','Secretaria','Municipio Radicado','Nombre Sede','Dane Sede','Ubicación',
               'Area','Area Otras','Antiguedad','Pregunta 1','Pregunta 1 Opcional','Pregunta 2','Pregunta 3','Pregunta 4','Pregunta 5','Pregunta 6',
               'Pregunta 6 Opcional','Pregunta 7','Pregunta 8','Pregunta 9','Pregunta 10','Pregunta 11','Pregunta 12','Pregunta 12 Opcional','Pregunta 13']

    formatos = ['General','0','General','General','General','General','General','General','General','General',
                'General','General','General','d/m/yy','0','General','General','General','0','General',
                'General','General','General','General','General','General','General','General','General','General',
                'General','General','General','General','General','General','General','General','General',]


    ancho_columnas =  [30,20,15,15,15,40,40,40,40,40,
                       40,40,40,40,40,40,40,40,40,40,
                       60,60,60,60,60,60,60,60,60,60,
                       60,60,60,60,60,60,60,60,60]

    contenidos = []

    for encuestado in PercepcionInicial.objects.all():
        contenidos.append([
            'ENC-'+unicode(encuestado.id),
            encuestado.docente_preinscrito.cedula,
            encuestado.docente_preinscrito.primer_apellido,
            encuestado.docente_preinscrito.segundo_apellido,
            encuestado.docente_preinscrito.primer_nombre,
            encuestado.docente_preinscrito.segundo_nombre,
            encuestado.docente_preinscrito.cargo,
            encuestado.docente_preinscrito.correo,
            encuestado.docente_preinscrito.telefono_fijo,
            encuestado.docente_preinscrito.telefono_celular,
            encuestado.docente_preinscrito.departamento.nombre if encuestado.docente_preinscrito.departamento != None else '',
            encuestado.docente_preinscrito.municipio.nombre if encuestado.docente_preinscrito.municipio != None else '',
            'SI' if encuestado.docente_preinscrito.verificado else 'NO',
            encuestado.docente_preinscrito.fecha,
            encuestado.docente_preinscrito.radicado.numero if encuestado.docente_preinscrito.radicado != None else '',
            encuestado.docente_preinscrito.radicado.secretaria.nombre if encuestado.docente_preinscrito.radicado.secretaria != None else '',
            encuestado.docente_preinscrito.radicado.municipio.nombre if encuestado.docente_preinscrito.radicado.municipio != None else '',
            encuestado.docente_preinscrito.radicado.nombre_sede if encuestado.docente_preinscrito.radicado != None else '',
            encuestado.docente_preinscrito.radicado.dane_sede if encuestado.docente_preinscrito.radicado != None else '',
            encuestado.docente_preinscrito.radicado.ubicacion if encuestado.docente_preinscrito.radicado != None else '',
            encuestado.area,
            encuestado.area_1,
            encuestado.antiguedad,
            encuestado.pregunta_1,
            encuestado.pregunta_1_1,
            encuestado.pregunta_2,
            encuestado.pregunta_3,
            encuestado.pregunta_4,
            encuestado.pregunta_5,
            encuestado.pregunta_6,
            encuestado.pregunta_6_1,
            encuestado.pregunta_7,
            encuestado.pregunta_8,
            encuestado.pregunta_9,
            encuestado.pregunta_10,
            encuestado.pregunta_11,
            encuestado.pregunta_12,
            encuestado.pregunta_12_1,
            encuestado.pregunta_13,
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def radicados(email):
    usuario = User.objects.get(email=email)
    nombre = "Base de datos radicados"
    proceso = "DB-INF01"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    titulos = ['ID','Numero','Municipio','Departamento','Secretaria','Nombre Sede','Dane Sede','Tipo','Ubicación']

    formatos = ['General','General','General','General','General','General','General','General','General']


    ancho_columnas =  [30,30,30,30,30,30,30,30,30]

    contenidos = []

    for radicado in Radicado.objects.exclude(oculto=True):
        contenidos.append([
            'RAD-'+unicode(radicado.id),
            radicado.numero,
            radicado.municipio.nombre if radicado.municipio != None else '',
            radicado.municipio.departamento.nombre if radicado.municipio.departamento != None else '',
            radicado.secretaria.nombre if radicado.secretaria != None else '',
            radicado.nombre_sede,
            radicado.dane_sede,
            radicado.tipo,
            radicado.ubicacion
        ])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def pagos_mensual(email):
    usuario = User.objects.get(email=email)
    nombre = "Reporte mensual de pago a formadores"
    proceso = "COR-INF02"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion
    cortes = Cortes.objects.all()
    fechas = []

    for corte in cortes:
        fecha_corte = '01/' + str(corte.mes) + '/' + str(corte.year)
        fecha_corte = datetime.datetime.strptime(fecha_corte,"%d/%m/%Y").date()
        fechas.append(fecha_corte)

    fechas = list(set(fechas))


    titulos = ['ID','Nombres','Apellidos','Cedula','Región',
               'Correo','Celular','Cargo','Banco','Tipo cuenta','Numero cuenta']


    formatos = ['General','General','General','General','General',
               'General','General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15,
                       40,20,20,20,20,20]


    for fecha_corte in fechas:
        titulos.append(fecha_corte.strftime("%B de %Y"))
        formatos.append('"$"#,##0_);("$"#,##0)')
        ancho_columnas.append(30)

    contenidos = []



    for formador in Formador.objects.all():
        row =[
                'FOR-'+unicode(formador.id),
                formador.nombres,
                formador.apellidos,
                formador.cedula,
                formador.get_region_string(),
                formador.correo_personal,
                formador.celular_personal,
                formador.cargo.nombre if formador.cargo != None else '',
                formador.banco.nombre if formador.banco != None else '',
                formador.tipo_cuenta,
                formador.numero_cuenta,
            ]
        for fecha_corte in fechas:
            valor = 0
            for corte in Cortes.objects.filter(mes = str(fecha_corte.month).zfill(2), year = fecha_corte.year):
                revisiones = Revision.objects.filter(corte = corte, formador_revision = formador)
                for revision in revisiones:
                    for producto in revision.productos.all():
                        valor += producto.cantidad * producto.valor_entregable.valor

            row.append(valor)

        contenidos.append(row)

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def zip_hv(email):

    if os.path.exists("C:\\Temp\\hv.zip"):
        os.remove("C:\\Temp\\hv.zip")

    usuario = User.objects.get(email=email)
    nombre = "Zip: Hojas de vida"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")

    zip = zipfile.ZipFile('C:\\Temp\\hv.zip',"w",allowZip64=True)

    for soporte in SoporteFormadores.objects.filter(tipo__id = 3).exclude(oculto = True):
        if str(soporte.archivo) != '':
            if os.path.exists(soporte.archivo.path):
                zip.write(soporte.archivo.path,soporte.formador.get_full_name()+'/'+os.path.basename(soporte.archivo.path))

    zip.close()
    informe.archivo = File(open('C:\\Temp\\hv.zip'))
    informe.save()

    shutil.copy('C:\\Temp\\hv.zip',informe.archivo.path)

    return "Zip creado HV"

@app.task
def zip_contrato(email):

    if os.path.exists("C:\\Temp\\contratos.zip"):
        os.remove("C:\\Temp\\contratos.zip")

    usuario = User.objects.get(email=email)
    nombre = "Zip: Contratos"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")


    zip = zipfile.ZipFile('C:\\Temp\\contratos.zip',"w",allowZip64=True)

    for soporte in SoporteFormadores.objects.filter(tipo__id = 10).exclude(oculto = True):
        if str(soporte.archivo) != '':
            if os.path.exists(soporte.archivo.path):
                zip.write(soporte.archivo.path,soporte.formador.get_full_name()+'/'+os.path.basename(soporte.archivo.path))

    zip.close()
    informe.archivo = File(open('C:\\Temp\\contratos.zip'))
    informe.save()

    shutil.copy('C:\\Temp\\contratos.zip',informe.archivo.path)

    return "Zip creado Contrato"

@app.task
def acumulado_tipo_1(email):
    usuario = User.objects.get(email=email)
    nombre = "Acumulado formadores tipo 1"
    proceso = "REV-INF01"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    contrato_1 = Contratos.objects.filter(cargo__nombre="Formador Tipo 1").get(nombre = 'Capacitación 1')
    entregables_1 = ValorEntregable.objects.filter(contrato = contrato_1)

    contrato_2 = Contratos.objects.filter(cargo__nombre="Formador Tipo 1").get(nombre = 'Capacitación 2')
    entregables_2 = ValorEntregable.objects.filter(contrato = contrato_2)


    titulos = ['ID','Nombres','Apellidos','Cedula','Región','Correo','Celular','Cargo','Profesión','Inicio contrato',
               'Fin contrato','Banco','Tipo cuenta','Numero cuenta','Eps','Pensión','Arl']



    formatos = ['General','General','General','General','General','General','General','General','General','d/m/yy',
               'd/m/yy','General','General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15,50,25,20,20,10,
                       10,20,20,20,20,20,20]


    if list(entregables_1.values_list('entregable__id',flat=True)) == list(entregables_2.values_list('entregable__id',flat=True)):
        for entregable in entregables_1:
            titulos.append(entregable.entregable.sesion.nivel.nombre + ' - ' +
                           entregable.entregable.sesion.nombre + ' - ID:' +
                           str(entregable.entregable.id))
            formatos.append('General')
            ancho_columnas.append(30)
    else:
        raise ValueError('Los contratos no tienen los mismos entregables')


    titulos.append('Valor pagado')
    formatos.append('"$"#,##0_);("$"#,##0)')
    ancho_columnas.append(30)


    contenidos = []

    for formador in Formador.objects.filter(cargo__nombre="Formador Tipo 1").exclude(oculto=True):

        cantidad_list = []
        valor = 0
        for entregable in entregables_1:
            productos = Revision.objects.filter(formador_revision = formador).filter(productos__valor_entregable__entregable__id = entregable.entregable.id)
            cantidad = productos.values_list('productos__cantidad',flat=True)
            cantidad_list.append(sum(cantidad))
            valor_entregable = 0

            for x in productos.values_list('productos__cantidad','productos__valor_entregable__valor'):
                valor_entregable += x[0]*x[1]

            valor += valor_entregable

        contenidos.append([
            'FOR-'+unicode(formador.id),
            formador.nombres,
            formador.apellidos,
            formador.cedula,
            formador.get_region_string(),
            formador.correo_personal,
            formador.celular_personal,
            formador.cargo.nombre if formador.cargo != None else '',
            formador.profesion,
            formador.fecha_contratacion,
            formador.fecha_terminacion,
            formador.banco.nombre if formador.banco != None else '',
            formador.tipo_cuenta,
            formador.numero_cuenta,
            formador.eps,
            formador.pension,
            formador.arl
        ] + cantidad_list + [valor])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def acumulado_tipo_2(email):
    usuario = User.objects.get(email=email)
    nombre = "Acumulado formadores tipo 2"
    proceso = "REV-INF02"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    contrato_1 = Contratos.objects.filter(cargo__nombre="Formador Tipo 2").get(nombre = 'Capacitación 1')
    entregables_1 = ValorEntregable.objects.filter(contrato = contrato_1)

    contrato_2 = Contratos.objects.filter(cargo__nombre="Formador Tipo 2").get(nombre = 'Capacitación 2')
    entregables_2 = ValorEntregable.objects.filter(contrato = contrato_2)


    titulos = ['ID','Nombres','Apellidos','Cedula','Región','Correo','Celular','Cargo','Profesión','Inicio contrato',
               'Fin contrato','Banco','Tipo cuenta','Numero cuenta','Eps','Pensión','Arl']



    formatos = ['General','General','General','General','General','General','General','General','General','d/m/yy',
               'd/m/yy','General','General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15,50,25,20,20,10,
                       10,20,20,20,20,20,20]


    if list(entregables_1.values_list('entregable__id',flat=True)) == list(entregables_2.values_list('entregable__id',flat=True)):
        for entregable in entregables_1:
            titulos.append(entregable.entregable.sesion.nivel.nombre + ' - ' +
                           entregable.entregable.sesion.nombre + ' - ID:' +
                           str(entregable.entregable.id))
            formatos.append('General')
            ancho_columnas.append(30)
    else:
        raise ValueError('Los contratos no tienen los mismos entregables')


    titulos.append('Valor pagado')
    formatos.append('"$"#,##0_);("$"#,##0)')
    ancho_columnas.append(30)


    contenidos = []

    for formador in Formador.objects.filter(cargo__nombre="Formador Tipo 2").exclude(oculto=True):

        cantidad_list = []
        valor = 0
        for entregable in entregables_1:
            productos = Revision.objects.filter(formador_revision = formador).filter(productos__valor_entregable__entregable__id = entregable.entregable.id)
            cantidad = productos.values_list('productos__cantidad',flat=True)
            cantidad_list.append(sum(cantidad))
            valor_entregable = 0

            for x in productos.values_list('productos__cantidad','productos__valor_entregable__valor'):
                valor_entregable += x[0]*x[1]

            valor += valor_entregable

        contenidos.append([
            'FOR-'+unicode(formador.id),
            formador.nombres,
            formador.apellidos,
            formador.cedula,
            formador.get_region_string(),
            formador.correo_personal,
            formador.celular_personal,
            formador.cargo.nombre if formador.cargo != None else '',
            formador.profesion,
            formador.fecha_contratacion,
            formador.fecha_terminacion,
            formador.banco.nombre if formador.banco != None else '',
            formador.tipo_cuenta,
            formador.numero_cuenta,
            formador.eps,
            formador.pension,
            formador.arl
        ] + cantidad_list + [valor])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def acumulado_tipo_3(email):
    usuario = User.objects.get(email=email)
    nombre = "Acumulado formadores tipo 3"
    proceso = "REV-INF03"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    contrato_1 = Contratos.objects.filter(cargo__nombre="Formador Tipo 3").get(nombre = 'Capacitación 1')
    entregables_1 = ValorEntregable.objects.filter(contrato = contrato_1)

    contrato_2 = Contratos.objects.filter(cargo__nombre="Formador Tipo 3").get(nombre = 'Capacitación 2')
    entregables_2 = ValorEntregable.objects.filter(contrato = contrato_2)


    titulos = ['ID','Nombres','Apellidos','Cedula','Región','Correo','Celular','Cargo','Profesión','Inicio contrato',
               'Fin contrato','Banco','Tipo cuenta','Numero cuenta','Eps','Pensión','Arl']



    formatos = ['General','General','General','General','General','General','General','General','General','d/m/yy',
               'd/m/yy','General','General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15,50,25,20,20,10,
                       10,20,20,20,20,20,20]


    if list(entregables_1.values_list('entregable__id',flat=True)) == list(entregables_2.values_list('entregable__id',flat=True)):
        for entregable in entregables_1:
            titulos.append(entregable.entregable.sesion.nivel.nombre + ' - ' +
                           entregable.entregable.sesion.nombre + ' - ID:' +
                           str(entregable.entregable.id))
            formatos.append('General')
            ancho_columnas.append(30)
    else:
        raise ValueError('Los contratos no tienen los mismos entregables')


    titulos.append('Valor pagado')
    formatos.append('"$"#,##0_);("$"#,##0)')
    ancho_columnas.append(30)


    contenidos = []

    for formador in Formador.objects.filter(cargo__nombre="Formador Tipo 3").exclude(oculto=True):

        cantidad_list = []
        valor = 0
        for entregable in entregables_1:
            productos = Revision.objects.filter(formador_revision = formador).filter(productos__valor_entregable__entregable__id = entregable.entregable.id)
            cantidad = productos.values_list('productos__cantidad',flat=True)
            cantidad_list.append(sum(cantidad))
            valor_entregable = 0

            for x in productos.values_list('productos__cantidad','productos__valor_entregable__valor'):
                valor_entregable += x[0]*x[1]

            valor += valor_entregable

        contenidos.append([
            'FOR-'+unicode(formador.id),
            formador.nombres,
            formador.apellidos,
            formador.cedula,
            formador.get_region_string(),
            formador.correo_personal,
            formador.celular_personal,
            formador.cargo.nombre if formador.cargo != None else '',
            formador.profesion,
            formador.fecha_contratacion,
            formador.fecha_terminacion,
            formador.banco.nombre if formador.banco != None else '',
            formador.tipo_cuenta,
            formador.numero_cuenta,
            formador.eps,
            formador.pension,
            formador.arl
        ] + cantidad_list + [valor])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def acumulado_tipo_4(email):
    usuario = User.objects.get(email=email)
    nombre = "Acumulado formadores tipo 4"
    proceso = "REV-INF04"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion

    contrato_1 = Contratos.objects.filter(cargo__nombre="Formador Tipo 4").get(nombre = 'Capacitación 1')
    entregables_1 = ValorEntregable.objects.filter(contrato = contrato_1)

    contrato_2 = Contratos.objects.filter(cargo__nombre="Formador Tipo 4").get(nombre = 'Capacitación 2')
    entregables_2 = ValorEntregable.objects.filter(contrato = contrato_2)


    titulos = ['ID','Nombres','Apellidos','Cedula','Región','Correo','Celular','Cargo','Profesión','Inicio contrato',
               'Fin contrato','Banco','Tipo cuenta','Numero cuenta','Eps','Pensión','Arl']



    formatos = ['General','General','General','General','General','General','General','General','General','d/m/yy',
               'd/m/yy','General','General','General','General','General','General']


    ancho_columnas =  [30,20,20,15,15,50,25,20,20,10,
                       10,20,20,20,20,20,20]


    if list(entregables_1.values_list('entregable__id',flat=True)) == list(entregables_2.values_list('entregable__id',flat=True)):
        for entregable in entregables_1:
            titulos.append(entregable.entregable.sesion.nivel.nombre + ' - ' +
                           entregable.entregable.sesion.nombre + ' - ID:' +
                           str(entregable.entregable.id))
            formatos.append('General')
            ancho_columnas.append(30)
    else:
        raise ValueError('Los contratos no tienen los mismos entregables')


    titulos.append('Valor pagado')
    formatos.append('"$"#,##0_);("$"#,##0)')
    ancho_columnas.append(30)


    contenidos = []

    for formador in Formador.objects.filter(cargo__nombre="Formador Tipo 4").exclude(oculto=True):

        cantidad_list = []
        valor = 0
        for entregable in entregables_1:
            productos = Revision.objects.filter(formador_revision = formador).filter(productos__valor_entregable__entregable__id = entregable.entregable.id)
            cantidad = productos.values_list('productos__cantidad',flat=True)
            cantidad_list.append(sum(cantidad))
            valor_entregable = 0

            for x in productos.values_list('productos__cantidad','productos__valor_entregable__valor'):
                valor_entregable += x[0]*x[1]

            valor += valor_entregable

        contenidos.append([
            'FOR-'+unicode(formador.id),
            formador.nombres,
            formador.apellidos,
            formador.cedula,
            formador.get_region_string(),
            formador.correo_personal,
            formador.celular_personal,
            formador.cargo.nombre if formador.cargo != None else '',
            formador.profesion,
            formador.fecha_contratacion,
            formador.fecha_terminacion,
            formador.banco.nombre if formador.banco != None else '',
            formador.tipo_cuenta,
            formador.numero_cuenta,
            formador.eps,
            formador.pension,
            formador.arl
        ] + cantidad_list + [valor])

    output = construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso)
    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def matriz_chequeo(email,id_diplomado):
    usuario = User.objects.get(email=email)

    if id_diplomado == '1':
        nombre = "Matriz lista de chequeo de productos InnovaTIC"
    elif id_diplomado == '2':
        nombre = "Matriz lista de chequeo de productos TecnoTIC"
    elif id_diplomado == '3':
        nombre = "Matriz lista de chequeo de productos DirecTIC"
    else:
        nombre = "Matriz lista de chequeo de productos EscuelaTIC"


    proceso = "REV-INF05"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion
    output = StringIO()
    dict_productos = []

    if id_diplomado == '1':
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Innovatic.xlsx')
        ws = wb.get_sheet_by_name('InnovaTIC')
        diplomado = Diplomado.objects.get(id = 1)
        i = 6

        dict_productos = [{'letter':'Y','id':8},
                          {'letter':'Z','id':9},
                          {'letter':'AA','id':10},
                          {'letter':'AC','id':11},
                          {'letter':'AF','id':12},
                          {'letter':'AG','id':13},
                          {'letter':'AK','id':14},
                          {'letter':'AL','id':15},
                          {'letter':'AN','id':16},
                          {'letter':'AR','id':17},
                          {'letter':'AS','id':18},
                          {'letter':'AT','id':19},
                          {'letter':'AD','id':20},
                          {'letter':'AI','id':21},
                          {'letter':'AJ','id':22},
                          {'letter':'AO','id':23},
                          {'letter':'AQ','id':24},
                          {'letter':'AU','id':25},
                          {'letter':'AV','id':26},
                          {'letter':'AW','id':27},
                          {'letter':'AX','id':28},
                          {'letter':'AY','id':29},
                          {'letter':'BB','id':30},
                          {'letter':'BC','id':31},
                          {'letter':'BD','id':32},
                          {'letter':'BG','id':33},
                          {'letter':'BH','id':34},
                          {'letter':'BJ','id':35},
                          {'letter':'BN','id':36},
                          {'letter':'BO','id':37},
                          {'letter':'BP','id':38},
                          {'letter':'AZ','id':39},
                          {'letter':'BA','id':40},
                          {'letter':'BF','id':41},
                          {'letter':'BL','id':42},
                          {'letter':'BM','id':43},
                          {'letter':'BQ','id':44},
                          {'letter':'BR','id':45},
                          {'letter':'BS','id':46},
                          {'letter':'BT','id':47},
                          {'letter':'BU','id':48},
                          {'letter':'BX','id':49},
                          {'letter':'BY','id':50},
                          {'letter':'BZ','id':51},
                          {'letter':'CB','id':52},
                          {'letter':'CC','id':53},
                          {'letter':'CD','id':54},
                          {'letter':'CG','id':55},
                          {'letter':'CH','id':56},
                          {'letter':'CI','id':57},
                          {'letter':'BV','id':58},
                          {'letter':'CA','id':59},
                          {'letter':'CE','id':60},
                          {'letter':'CJ','id':61},
                          {'letter':'CK','id':62},
                          {'letter':'CL','id':63},
                          {'letter':'CM','id':64},
                          {'letter':'CO','id':65},
                          {'letter':'CR','id':66},
                          {'letter':'CS','id':67},
                          {'letter':'CN','id':68},
                          {'letter':'CQ','id':69},
                          ]

    elif id_diplomado == '2':
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Tecnotic.xlsx')
        ws = wb.get_sheet_by_name('TecnoTIC')
        diplomado = Diplomado.objects.get(id = 2)
        i = 6
    elif id_diplomado == '3':
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Directic.xlsx')
        ws = wb.get_sheet_by_name('DirecTIC')
        diplomado = Diplomado.objects.get(id = 3)
        i = 6
    else:
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Innovatic.xlsx')
        ws = wb.get_sheet_by_name('InnovaTIC')
        diplomado = Diplomado.objects.get(id = 4)
        i = 6

    number = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='right',vertical='center',wrap_text=False),
                   number_format='0',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
             )

    text = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='left',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                 )


    validado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                   fill=PatternFill(fill_type='solid',start_color='FF00B050',end_color='FF00B050')
                 )

    enviado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                   fill=PatternFill(fill_type='solid',start_color='FFFFC000',end_color='FFFFC000')
                 )

    cargado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                 )


    rechazado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                   fill=PatternFill(fill_type='solid',start_color='FFFF0000',end_color='FFFF0000')
                 )

    for beneficiario in Beneficiario.objects.filter(diplomado = diplomado).order_by('formador'):
        ws.cell('A'+str(i)).value = beneficiario.region.nombre.upper()
        ws.cell('A'+str(i)).style = text

        ws.cell('B'+str(i)).value = beneficiario.radicado.municipio.departamento.nombre.upper() if beneficiario.radicado != None else ''
        ws.cell('B'+str(i)).style = text

        ws.cell('C'+str(i)).value = beneficiario.radicado.secretaria.nombre.upper() if beneficiario.radicado != None else ''
        ws.cell('C'+str(i)).style = text

        ws.cell('D'+str(i)).value = beneficiario.radicado.numero if beneficiario.radicado != None else ''
        ws.cell('D'+str(i)).style = number

        ws.cell('E'+str(i)).value = ''
        ws.cell('E'+str(i)).style = text

        ws.cell('F'+str(i)).value = ''
        ws.cell('F'+str(i)).style = text

        ws.cell('G'+str(i)).value = beneficiario.radicado.dane_sede if beneficiario.radicado != None else ''
        ws.cell('G'+str(i)).style = number

        ws.cell('H'+str(i)).value = beneficiario.radicado.nombre_sede.upper() if beneficiario.radicado != None else ''
        ws.cell('H'+str(i)).style = text

        ws.cell('I'+str(i)).value = beneficiario.radicado.municipio.nombre.upper() if beneficiario.radicado != None else ''
        ws.cell('I'+str(i)).style = text

        ws.cell('J'+str(i)).value = beneficiario.radicado.ubicacion if beneficiario.radicado != None else ''
        ws.cell('J'+str(i)).style = text

        ws.cell('K'+str(i)).value = beneficiario.get_grupo().upper()
        ws.cell('K'+str(i)).style = text

        ws.cell('L'+str(i)).value = beneficiario.formador.get_full_name().upper()
        ws.cell('L'+str(i)).style = text

        ws.cell('M'+str(i)).value = beneficiario.formador.cedula
        ws.cell('M'+str(i)).style = number

        ws.cell('N'+str(i)).value = beneficiario.apellidos.upper()
        ws.cell('N'+str(i)).style = text

        ws.cell('O'+str(i)).value = beneficiario.nombres.upper()
        ws.cell('O'+str(i)).style = text

        ws.cell('P'+str(i)).value = beneficiario.cedula
        ws.cell('P'+str(i)).style = number

        ws.cell('Q'+str(i)).value = beneficiario.correo
        ws.cell('Q'+str(i)).style = text

        ws.cell('R'+str(i)).value = beneficiario.telefono_fijo
        ws.cell('R'+str(i)).style = text

        ws.cell('S'+str(i)).value = beneficiario.telefono_celular
        ws.cell('S'+str(i)).style = text

        ws.cell('T'+str(i)).value = beneficiario.area.nombre.upper() if beneficiario.area != None else ''
        ws.cell('T'+str(i)).style = text

        ws.cell('U'+str(i)).value = beneficiario.grado.nombre.upper() if beneficiario.grado != None else ''
        ws.cell('U'+str(i)).style = text

        ws.cell('V'+str(i)).value = ''
        ws.cell('V'+str(i)).style = text

        ws.cell('W'+str(i)).value = beneficiario.genero
        ws.cell('W'+str(i)).style = text

        ws.cell('X'+str(i)).value = beneficiario.estado
        ws.cell('X'+str(i)).style = text

        ws.cell('Y'+str(i)).value = ''
        ws.cell('Y'+str(i)).style = text

        for producto in dict_productos:
            entregable = Entregable.objects.get(id = producto['id'])

            evidencias_cargado = Evidencia.objects.filter(beneficiarios_cargados = beneficiario,entregable = entregable)
            evidencias_validado = Evidencia.objects.filter(beneficiarios_validados = beneficiario,entregable = entregable)

            q = Q(evidencias__id__in = evidencias_cargado.values_list('id',flat=True)) | \
                Q(evidencias__id__in = evidencias_validado.values_list('id',flat=True))

            reds = Red.objects.filter(q)

            if evidencias_validado.count() > 0:
                red = Red.objects.get(evidencias__id = evidencias_validado[0].id)
                ws.cell(producto['letter']+str(i)).value = 'RED-' + str(red.id)
                ws.cell(producto['letter']+str(i)).style = validado
                ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + evidencias_validado[0].get_archivo_url()
                ws.cell( producto['letter'] + str(i) ).comment = Comment('SIC-' + str(evidencias_validado[0].id),'SICAN')


            elif evidencias_validado.count() == 0 and evidencias_cargado.count() > 0:
                ultima_evidencia_cargada = evidencias_cargado[len(evidencias_cargado)-1]

                try:
                    red = reds.get(evidencias__id = ultima_evidencia_cargada.id)
                except:
                    # si ningun red contiene la evidencia
                    ws.cell(producto['letter']+str(i)).value = 'SIC-' + str(ultima_evidencia_cargada.id)
                    ws.cell(producto['letter']+str(i)).style = cargado
                    ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + ultima_evidencia_cargada.get_archivo_url()

                else:
                    # si hay un red que contiene la evidencia
                    if red.retroalimentacion:
                        # si el red fue retroalimentado

                        evidencias_rechazado = Evidencia.objects.filter(beneficiarios_rechazados__beneficiario_rechazo = beneficiario,
                                                                        beneficiarios_rechazados__red_id = red.id,
                                                                        beneficiarios_rechazados__evidencia_id = ultima_evidencia_cargada.id)

                        if evidencias_rechazado > 0:
                            try:
                                causa = evidencias_rechazado[0].beneficiarios_rechazados.get(beneficiario_rechazo = beneficiario).observacion
                            except:
                                causa = ''
                            ws.cell(producto['letter']+str(i)).value = 'RED-' + str(red.id)
                            ws.cell(producto['letter']+str(i)).style = rechazado
                            ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + ultima_evidencia_cargada.get_archivo_url()
                            ws.cell( producto['letter'] + str(i) ).comment = Comment('SIC-' + str(ultima_evidencia_cargada.id) + ':\n' + causa,'SICAN')
                        else:
                            pass


                    else:
                        #si el red no ha sido retroalimentado
                        ws.cell(producto['letter']+str(i)).value = 'RED-' + str(red.id)
                        ws.cell(producto['letter']+str(i)).style = enviado
                        ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + ultima_evidencia_cargada.get_archivo_url()
                        ws.cell( producto['letter'] + str(i) ).comment = Comment('SIC-' + str(ultima_evidencia_cargada.id),'SICAN')


            else:
                ws.cell(producto['letter']+str(i)).style = text
        i += 1

    wb.save(output)

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"

@app.task
def matriz_chequeo_formador(email,id_formador):
    usuario = User.objects.get(email=email)

    formador = Formador.objects.get(id = int(id_formador))

    if formador.cargo.nombre == 'Formador Tipo 1':
        id_diplomado = '1'
    elif formador.cargo.nombre == 'Formador Tipo 2':
        id_diplomado = '2'
    elif formador.cargo.nombre == 'Formador Tipo 3':
        id_diplomado = '3'
    else:
        id_diplomado = '4'


    if id_diplomado == '1':
        nombre = "Matriz lista de chequeo de productos InnovaTIC - " + formador.get_full_name()
    elif id_diplomado == '2':
        nombre = "Matriz lista de chequeo de productos TecnoTIC - " + formador.get_full_name()
    elif id_diplomado == '3':
        nombre = "Matriz lista de chequeo de productos DirecTIC - " + formador.get_full_name()
    else:
        nombre = "Matriz lista de chequeo de productos EscuelaTIC - " + formador.get_full_name()


    proceso = "REV-INF06"
    informe = InformesExcel.objects.create(usuario = usuario,nombre=nombre,progreso="0%")
    fecha = informe.creacion
    output = StringIO()
    dict_productos = []

    if id_diplomado == '1':
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Innovatic.xlsx')
        ws = wb.get_sheet_by_name('InnovaTIC')
        diplomado = Diplomado.objects.get(id = 1)
        i = 6

        dict_productos = [{'letter':'Y','id':8},
                          {'letter':'Z','id':9},
                          {'letter':'AA','id':10},
                          {'letter':'AC','id':11},
                          {'letter':'AF','id':12},
                          {'letter':'AG','id':13},
                          {'letter':'AK','id':14},
                          {'letter':'AL','id':15},
                          {'letter':'AN','id':16},
                          {'letter':'AR','id':17},
                          {'letter':'AS','id':18},
                          {'letter':'AT','id':19},
                          {'letter':'AD','id':20},
                          {'letter':'AI','id':21},
                          {'letter':'AJ','id':22},
                          {'letter':'AO','id':23},
                          {'letter':'AQ','id':24},
                          {'letter':'AU','id':25},
                          {'letter':'AV','id':26},
                          {'letter':'AW','id':27},
                          {'letter':'AX','id':28},
                          {'letter':'AY','id':29},
                          {'letter':'BB','id':30},
                          {'letter':'BC','id':31},
                          {'letter':'BD','id':32},
                          {'letter':'BG','id':33},
                          {'letter':'BH','id':34},
                          {'letter':'BJ','id':35},
                          {'letter':'BN','id':36},
                          {'letter':'BO','id':37},
                          {'letter':'BP','id':38},
                          {'letter':'AZ','id':39},
                          {'letter':'BA','id':40},
                          {'letter':'BF','id':41},
                          {'letter':'BL','id':42},
                          {'letter':'BM','id':43},
                          {'letter':'BQ','id':44},
                          {'letter':'BR','id':45},
                          {'letter':'BS','id':46},
                          {'letter':'BT','id':47},
                          {'letter':'BU','id':48},
                          {'letter':'BX','id':49},
                          {'letter':'BY','id':50},
                          {'letter':'BZ','id':51},
                          {'letter':'CB','id':52},
                          {'letter':'CC','id':53},
                          {'letter':'CD','id':54},
                          {'letter':'CG','id':55},
                          {'letter':'CH','id':56},
                          {'letter':'CI','id':57},
                          {'letter':'BV','id':58},
                          {'letter':'CA','id':59},
                          {'letter':'CE','id':60},
                          {'letter':'CJ','id':61},
                          {'letter':'CK','id':62},
                          {'letter':'CL','id':63},
                          {'letter':'CM','id':64},
                          {'letter':'CO','id':65},
                          {'letter':'CR','id':66},
                          {'letter':'CS','id':67},
                          {'letter':'CN','id':68},
                          {'letter':'CQ','id':69},
                          ]

    elif id_diplomado == '2':
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Tecnotic.xlsx')
        ws = wb.get_sheet_by_name('TecnoTIC')
        diplomado = Diplomado.objects.get(id = 2)
        i = 6
    elif id_diplomado == '3':
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Directic.xlsx')
        ws = wb.get_sheet_by_name('DirecTIC')
        diplomado = Diplomado.objects.get(id = 3)
        i = 6
    else:
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Matriz Innovatic.xlsx')
        ws = wb.get_sheet_by_name('InnovaTIC')
        diplomado = Diplomado.objects.get(id = 4)
        i = 6

    number = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='right',vertical='center',wrap_text=False),
                   number_format='0',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
             )

    text = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='left',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                 )


    validado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                   fill=PatternFill(fill_type='solid',start_color='FF00B050',end_color='FF00B050')
                 )

    enviado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                   fill=PatternFill(fill_type='solid',start_color='FFFFC000',end_color='FFFFC000')
                 )

    cargado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                 )


    rechazado = Style(font=Font(name='Calibri',size=12),
                   alignment=Alignment(horizontal='center',vertical='center',wrap_text=False),
                   number_format='General',
                   border=Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin')),
                   fill=PatternFill(fill_type='solid',start_color='FFFF0000',end_color='FFFF0000')
                 )

    for beneficiario in Beneficiario.objects.filter(diplomado = diplomado).order_by('formador').filter(formador__id = id_formador):
        ws.cell('A'+str(i)).value = beneficiario.region.nombre.upper()
        ws.cell('A'+str(i)).style = text

        ws.cell('B'+str(i)).value = beneficiario.radicado.municipio.departamento.nombre.upper() if beneficiario.radicado != None else ''
        ws.cell('B'+str(i)).style = text

        ws.cell('C'+str(i)).value = beneficiario.radicado.secretaria.nombre.upper() if beneficiario.radicado != None else ''
        ws.cell('C'+str(i)).style = text

        ws.cell('D'+str(i)).value = beneficiario.radicado.numero if beneficiario.radicado != None else ''
        ws.cell('D'+str(i)).style = number

        ws.cell('E'+str(i)).value = ''
        ws.cell('E'+str(i)).style = text

        ws.cell('F'+str(i)).value = ''
        ws.cell('F'+str(i)).style = text

        ws.cell('G'+str(i)).value = beneficiario.radicado.dane_sede if beneficiario.radicado != None else ''
        ws.cell('G'+str(i)).style = number

        ws.cell('H'+str(i)).value = beneficiario.radicado.nombre_sede.upper() if beneficiario.radicado != None else ''
        ws.cell('H'+str(i)).style = text

        ws.cell('I'+str(i)).value = beneficiario.radicado.municipio.nombre.upper() if beneficiario.radicado != None else ''
        ws.cell('I'+str(i)).style = text

        ws.cell('J'+str(i)).value = beneficiario.radicado.ubicacion if beneficiario.radicado != None else ''
        ws.cell('J'+str(i)).style = text

        ws.cell('K'+str(i)).value = beneficiario.get_grupo().upper()
        ws.cell('K'+str(i)).style = text

        ws.cell('L'+str(i)).value = beneficiario.formador.get_full_name().upper()
        ws.cell('L'+str(i)).style = text

        ws.cell('M'+str(i)).value = beneficiario.formador.cedula
        ws.cell('M'+str(i)).style = number

        ws.cell('N'+str(i)).value = beneficiario.apellidos.upper()
        ws.cell('N'+str(i)).style = text

        ws.cell('O'+str(i)).value = beneficiario.nombres.upper()
        ws.cell('O'+str(i)).style = text

        ws.cell('P'+str(i)).value = beneficiario.cedula
        ws.cell('P'+str(i)).style = number

        ws.cell('Q'+str(i)).value = beneficiario.correo
        ws.cell('Q'+str(i)).style = text

        ws.cell('R'+str(i)).value = beneficiario.telefono_fijo
        ws.cell('R'+str(i)).style = text

        ws.cell('S'+str(i)).value = beneficiario.telefono_celular
        ws.cell('S'+str(i)).style = text

        ws.cell('T'+str(i)).value = beneficiario.area.nombre.upper() if beneficiario.area != None else ''
        ws.cell('T'+str(i)).style = text

        ws.cell('U'+str(i)).value = beneficiario.grado.nombre.upper() if beneficiario.grado != None else ''
        ws.cell('U'+str(i)).style = text

        ws.cell('V'+str(i)).value = ''
        ws.cell('V'+str(i)).style = text

        ws.cell('W'+str(i)).value = beneficiario.genero
        ws.cell('W'+str(i)).style = text

        ws.cell('X'+str(i)).value = beneficiario.estado
        ws.cell('X'+str(i)).style = text

        ws.cell('Y'+str(i)).value = ''
        ws.cell('Y'+str(i)).style = text

        for producto in dict_productos:
            entregable = Entregable.objects.get(id = producto['id'])

            evidencias_cargado = Evidencia.objects.filter(beneficiarios_cargados = beneficiario,entregable = entregable)
            evidencias_validado = Evidencia.objects.filter(beneficiarios_validados = beneficiario,entregable = entregable)

            q = Q(evidencias__id__in = evidencias_cargado.values_list('id',flat=True)) | \
                Q(evidencias__id__in = evidencias_validado.values_list('id',flat=True))

            reds = Red.objects.filter(q)

            if evidencias_validado.count() > 0:
                red = Red.objects.get(evidencias__id = evidencias_validado[0].id)
                ws.cell(producto['letter']+str(i)).value = 'RED-' + str(red.id)
                ws.cell(producto['letter']+str(i)).style = validado
                ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + evidencias_validado[0].get_archivo_url()
                ws.cell( producto['letter'] + str(i) ).comment = Comment('SIC-' + str(evidencias_validado[0].id),'SICAN')


            elif evidencias_validado.count() == 0 and evidencias_cargado.count() > 0:
                ultima_evidencia_cargada = evidencias_cargado[len(evidencias_cargado)-1]

                try:
                    red = reds.get(evidencias__id = ultima_evidencia_cargada.id)
                except:
                    # si ningun red contiene la evidencia
                    ws.cell(producto['letter']+str(i)).value = 'SIC-' + str(ultima_evidencia_cargada.id)
                    ws.cell(producto['letter']+str(i)).style = cargado
                    ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + ultima_evidencia_cargada.get_archivo_url()

                else:
                    # si hay un red que contiene la evidencia
                    if red.retroalimentacion:
                        # si el red fue retroalimentado

                        evidencias_rechazado = Evidencia.objects.filter(beneficiarios_rechazados__beneficiario_rechazo = beneficiario,
                                                                        beneficiarios_rechazados__red_id = red.id,
                                                                        beneficiarios_rechazados__evidencia_id = ultima_evidencia_cargada.id)

                        if evidencias_rechazado > 0:
                            try:
                                causa = evidencias_rechazado[0].beneficiarios_rechazados.get(beneficiario_rechazo = beneficiario).observacion
                            except:
                                causa = ''
                            ws.cell(producto['letter']+str(i)).value = 'RED-' + str(red.id)
                            ws.cell(producto['letter']+str(i)).style = rechazado
                            ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + ultima_evidencia_cargada.get_archivo_url()
                            ws.cell( producto['letter'] + str(i) ).comment = Comment('SIC-' + str(ultima_evidencia_cargada.id) + ':\n' + causa,'SICAN')
                        else:
                            pass


                    else:
                        #si el red no ha sido retroalimentado
                        ws.cell(producto['letter']+str(i)).value = 'RED-' + str(red.id)
                        ws.cell(producto['letter']+str(i)).style = enviado
                        ws.cell( producto['letter'] + str(i) ).hyperlink = 'https://sican.asoandes.org' + ultima_evidencia_cargada.get_archivo_url()
                        ws.cell( producto['letter'] + str(i) ).comment = Comment('SIC-' + str(ultima_evidencia_cargada.id),'SICAN')


            else:
                ws.cell(producto['letter']+str(i)).style = text
        i += 1

    wb.save(output)

    filename = unicode(informe.creacion) + '.xlsx'
    informe.archivo.save(filename,File(output))
    return "Reporte generado exitosamente"