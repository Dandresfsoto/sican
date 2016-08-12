from __future__ import unicode_literals
import openpyxl
from StringIO import StringIO
from sican.settings import base as settings
from openpyxl.drawing.image import Image
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font

def construir_reporte(titulos,contenidos,formatos,ancho_columnas,nombre,fecha,usuario,proceso):
    if len(ancho_columnas) != len(formatos) != len(titulos) != len(contenidos[0]):
        raise Exception('El arreglo de filas y columnas tienen distinta longitud')
    else:
        output = StringIO()
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Formato_informe.xlsx')
        ws = wb.get_sheet_by_name('Hoja1')
        logo_andes = Image(settings.STATICFILES_DIRS[0]+'/img/andes_logo_informe.png',size=(145,145))
        logo_andes.drawing.top = 17
        logo_andes.drawing.left = 8
        ws.add_image(logo_andes)
        ws.cell('B1').value = "   Nombre: " + nombre
        ws.cell('B3').value = "   Fecha: " + fecha.strftime("%A %d de %B a las %X %p")
        ws.cell('B5').value = "   Usuario: " + usuario.email
        ws.cell('B7').value = "   Proceso: " + proceso


        row_num = 9
        for col_num in xrange(len(titulos)):
            ws.cell(row=row_num, column=col_num+1).value = titulos[col_num]
            if col_num != 0:
                ws.column_dimensions[openpyxl.cell.get_column_letter(col_num+1)].width = ancho_columnas[col_num]

        for contenido in contenidos:
            row_num += 1
            for col_num in xrange(len(contenido)):
                if contenido[col_num] == True:
                    ws.cell(row=row_num,column=col_num+1).value = "SI"
                if contenido[col_num] == False:
                    ws.cell(row=row_num,column=col_num+1).value = "NO"
                if contenido[col_num] == None:
                    ws.cell(row=row_num,column=col_num+1).value = ""
                else:
                    ws.cell(row=row_num,column=col_num+1).value = contenido[col_num]

                ws.cell(row=row_num,column=col_num+1).style = Style(font=Font(name='Arial',size=10),
                                                                        alignment=Alignment(
                                                                            horizontal='center',
                                                                            vertical='center',
                                                                            wrap_text=True
                                                                        ),
                                                                    number_format=formatos[col_num]
                                                                    )

        wb.save(output)
        return output