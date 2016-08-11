#!/usr/bin/env python
# -*- coding: utf-8 -*-
from __future__ import unicode_literals
from __future__ import absolute_import

from formadores.models import SolicitudTransporte
import openpyxl
from sican.settings import base as settings
from sican.celery import app
from StringIO import StringIO
from django.core.files import File
from openpyxl.drawing.image import Image
from datetime import datetime


@app.task
def construir_pdf(id_solicitud):
    solicitud = SolicitudTransporte.objects.get(id = id_solicitud)
    desplazamientos = solicitud.desplazamientos.all()

    output = StringIO()

    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Formato de transporte.xlsx')
    ws = wb.get_sheet_by_name('Hoja1')

    logo_cpe = Image(settings.STATICFILES_DIRS[0]+'/img/cpe_transporte.png',size=(120,120))
    logo_cpe.drawing.top = 22
    logo_cpe.drawing.left = 56

    logo_pais = Image(settings.STATICFILES_DIRS[0]+'/img/pais_transporte.png',size=(194,142))
    logo_pais.drawing.top = 22
    logo_pais.drawing.left = 1242

    ws.add_image(logo_cpe)
    ws.add_image(logo_pais)

    departamentos = desplazamientos.values_list('departamento_origen__nombre',flat=True).distinct()
    departamentos_unicode = ''

    for departamento in departamentos:
        departamentos_unicode += departamento + ', '

    municipios = desplazamientos.values_list('municipio_origen__nombre',flat=True).distinct()
    municipios_unicode = ''

    for municipio in municipios:
        municipios_unicode += municipio + ', '

    ws.cell('B6').value = 'REGIÓN: '+solicitud.formador.get_region_string()
    ws.cell('B9').value = 'NOMBRE DEL FORMADOR Y/O INTEGRANTE DEL EQUIPO: '+solicitud.formador.get_full_name()
    ws.cell('B10').value = 'DEPARTAMENTO: '+departamentos_unicode[:-2]
    ws.cell('B11').value = 'MUNICIPIOS: '+municipios_unicode[:-2]

    i = 15
    for desplazamiento in desplazamientos:
        ws.cell(row=i,column=2).value = desplazamiento.fecha.strftime('%d/%m/%Y')
        ws.cell(row=i,column=3).value = solicitud.formador.get_full_name()
        ws.cell(row=i,column=5).value = solicitud.formador.celular_personal
        ws.cell(row=i,column=6).value = desplazamiento.departamento_origen.nombre + ', ' + desplazamiento.municipio_origen.nombre
        ws.cell(row=i,column=7).value = desplazamiento.departamento_destino.nombre + ', ' + desplazamiento.municipio_destino.nombre
        ws.cell(row=i,column=8).value = desplazamiento.valor
        i += 1

    wb.save(output)
    filename = unicode(solicitud.creacion) + '.xlsx'
    solicitud.pdf.save(filename,File(output))
    return id_solicitud