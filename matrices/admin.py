from django.contrib import admin
from matrices.models import Area, Grado, Beneficiario
from matrices.models import CargaMasiva
from formadores.models import Grupos, Formador
import openpyxl
from sican.settings import base as settings
from administrativos.models import Administrativo
# Register your models here.

admin.site.register(Area)
admin.site.register(Grado)

def grupos_colombia_aprende(modeladmin, request, queryset):
    beneficiarios = Beneficiario.objects.exclude(usuario_colombia_aprende = '').exclude(radicado = None)
    grupos_lista_id = beneficiarios.values_list('grupo__id',flat=True).distinct()
    grupos = Grupos.objects.filter(id__in = grupos_lista_id).exclude(formador__usuario_colombia_aprende = '')
    formadores_lista_id = grupos.values_list('formador__id',flat=True).distinct()
    formadores = Formador.objects.filter(id__in = formadores_lista_id)

    for grupo in grupos:
        wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Formato GD.xlsx')
        ws = wb.get_sheet_by_name('Hoja1')

        fila = 2
        for beneficiario in Beneficiario.objects.filter(id__in = beneficiarios.values_list('id',flat=True)).filter(grupo = grupo):
            ws['A'+str(fila)] = 'ASOANDES'
            ws['B'+str(fila)] = beneficiario.region.numero
            ws['C'+str(fila)] = beneficiario.radicado.municipio.departamento.nombre
            ws['D'+str(fila)] = beneficiario.radicado.municipio.nombre
            ws['E'+str(fila)] = Administrativo.objects.get(correo_corporativo = beneficiario.formador.lider.email).cedula
            ws['F'+str(fila)] = Administrativo.objects.get(correo_corporativo = beneficiario.formador.lider.email).nombres
            ws['G'+str(fila)] = Administrativo.objects.get(correo_corporativo = beneficiario.formador.lider.email).apellidos
            ws['H'+str(fila)] = beneficiario.formador.lider.email
            ws['I'+str(fila)] = Administrativo.objects.get(correo_corporativo = beneficiario.formador.lider.email).celular_personal
            ws['J'+str(fila)] = Administrativo.objects.get(correo_corporativo = beneficiario.formador.lider.email).usuario_colombia_aprende
            ws['K'+str(fila)] = beneficiario.formador.cedula
            ws['L'+str(fila)] = beneficiario.formador.nombres
            ws['M'+str(fila)] = beneficiario.formador.apellidos
            ws['N'+str(fila)] = beneficiario.formador.correo_personal
            ws['O'+str(fila)] = beneficiario.formador.celular_personal
            ws['P'+str(fila)] = beneficiario.formador.usuario_colombia_aprende
            ws['Q'+str(fila)] = beneficiario.diplomado.nombre
            ws['R'+str(fila)] = beneficiario.cedula
            ws['S'+str(fila)] = beneficiario.nombres
            ws['T'+str(fila)] = beneficiario.apellidos
            ws['U'+str(fila)] = beneficiario.correo
            ws['V'+str(fila)] = beneficiario.telefono_celular
            ws['W'+str(fila)] = beneficiario.usuario_colombia_aprende
            fila += 1


        wb.save('C:\\Temp\\Colombia\\'+grupo.formador.codigo_ruta + '-' + grupo.nombre + '.xlsx')
grupos_colombia_aprende.short_description = 'Usuarios colombia aprende'

class BeneficiarioAdmin(admin.ModelAdmin):
    list_display = ['nombres','apellidos']
    ordering = ['nombres']
    actions = [grupos_colombia_aprende]

admin.site.register(Beneficiario, BeneficiarioAdmin)
admin.site.register(CargaMasiva)