from django.contrib import admin
from formadores.models import Formador
from formadores.models import SolicitudTransporte, Desplazamiento
from formadores.models import Producto, Revision, Cortes
from financiera.tasks import cortes
import openpyxl
from sican.settings import base as settings
from administrativos.models import Administrativo

# Register your models here.


def grupos_colombia_aprende(modeladmin, request, queryset):

    wb = openpyxl.load_workbook(filename=settings.STATICFILES_DIRS[0]+'/documentos/Formato GD.xlsx')
    ws = wb.get_sheet_by_name('Hoja1')
    fila = 2

    for formador in Formador.objects.exclude(usuario_colombia_aprende = '').exclude(lider = None).exclude(lider__id = 1):

        diplomado = ''

        if formador.cargo.nombre == 'Formador Tipo 1':
            diplomado = 'INNOVATIC'
        if formador.cargo.nombre == 'Formador Tipo 2':
            diplomado = 'TECNOTIC'
        if formador.cargo.nombre == 'Formador Tipo 3':
            diplomado = 'DIRECTIC'
        if formador.cargo.nombre == 'Formador Tipo 4':
            diplomado = 'ESCUELA TIC'

        for i in range(1,5):

            ws['A'+str(fila)] = 'ASOANDES'
            ws['B'+str(fila)] = formador.get_region_string()
            ws['C'+str(fila)] = ''
            ws['D'+str(fila)] = ''
            ws['E'+str(fila)] = Administrativo.objects.exclude(oculto=True).get(correo_corporativo = formador.lider.email).cedula
            ws['F'+str(fila)] = Administrativo.objects.exclude(oculto=True).get(correo_corporativo = formador.lider.email).nombres
            ws['G'+str(fila)] = Administrativo.objects.exclude(oculto=True).get(correo_corporativo = formador.lider.email).apellidos
            ws['H'+str(fila)] = formador.lider.email
            ws['I'+str(fila)] = Administrativo.objects.exclude(oculto=True).get(correo_corporativo = formador.lider.email).celular_personal
            ws['J'+str(fila)] = Administrativo.objects.exclude(oculto=True).get(correo_corporativo = formador.lider.email).usuario_colombia_aprende
            ws['K'+str(fila)] = formador.cedula
            ws['L'+str(fila)] = formador.nombres
            ws['M'+str(fila)] = formador.apellidos
            ws['N'+str(fila)] = formador.correo_personal
            ws['O'+str(fila)] = formador.celular_personal
            ws['P'+str(fila)] = formador.usuario_colombia_aprende
            ws['Q'+str(fila)] = diplomado + '-' + str(i)
            ws['R'+str(fila)] = ''
            ws['S'+str(fila)] = ''
            ws['T'+str(fila)] = ''
            ws['U'+str(fila)] = ''
            ws['V'+str(fila)] = ''
            ws['W'+str(fila)] = ''
            fila += 1

    wb.save('C:\\Temp\\Colombia\\Formadores.xlsx')

grupos_colombia_aprende.short_description = 'Usuarios colombia aprende'

class FormadorAdmin(admin.ModelAdmin):
    list_display = ['get_full_name','cedula']
    ordering = ['cedula']
    actions = [grupos_colombia_aprende]

admin.site.register(Formador,FormadorAdmin)
admin.site.register(SolicitudTransporte)
admin.site.register(Desplazamiento)
admin.site.register(Producto)
admin.site.register(Revision)


def cortes_archivo(modeladmin, request, queryset):
   for obj in queryset:
       cortes.delay(request.user.email,obj.id)
cortes_archivo.short_description = 'Actualizar archivo de corte'


class CortesAdmin(admin.ModelAdmin):
    list_display = ['descripcion']
    ordering = ['id']
    actions = [cortes_archivo]

admin.site.register(Cortes,CortesAdmin)